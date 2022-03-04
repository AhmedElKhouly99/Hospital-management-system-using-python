import openpyxl

wb = openpyxl.load_workbook('hospital.xlsx')

def Read(a, b, counter, c = 1):
    
    sheet = wb[a]
    asheet = wb.active
    #data_num = sheet.max_column
    #data = sheet.
    if c == 1:
        
        r = 2
        while(sheet.max_row>= r):
            if sheet.cell(row=r, column=1).value != None:
                b[str(sheet.cell(row=r, column=1).value)] = list()
                c = 2
                while sheet.max_column >= c:
                    b[sheet['A'+str(r)].value].append(sheet.cell(row=r, column=c).value)
                    c+=1
                r+=1
            else:
                break
                
        counter = sheet.cell(row=2, column=6).value
        #b=dict()

        #j=2
        #for row in sheet.iter_rows(min_row=1, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
        #    if sheet.cell(row=j, column=1).value != None:
        #        b[str(sheet.cell(row=j, column=1).value)] = list()
        #        for cell in row:
        #            if cell != None:
        #                b[sheet['A'+str(j)].value].append(cell.value)
        #                #print(cell.value)
        #            else :
        #                break
        #        j +=1
        #    else:
        #        break
            
        #for x in range(1, sheet.max_row+1):
        #    b[sheet.cell(row=x, column=1).value] = str(sheet.cell(row=x, column=2).value)
    else:
        i = 0
        r = 2
        while sheet.max_row >= r:
            b[i] = sheet.cell(row=r, column=3).value
            i+=1
            
        #i = 0
        #for x in range(1, sheet.max_row+1):
        #    b[i] = sheet.cell(row=x, column=1).value
        #    i+=1 
    wb.save("hospital.xlsx")        
            
def Write(a, b, counter, c = 1): 
    
    sheet = wb[a]
    asheet = wb.active
    #j=1
    
    
    
    r = 2
    while(sheet.max_row>= r):
        c = 1
        while sheet.max_column >= c:
            sheet.cell(row=r, column=c).value = None
            c+=1
        r+=1
    
    if c == 1:
        ro = 2        
        for item in b:
            listb = b[item] 
            size  = len(b[item])
            sheet.cell(row=ro, column=1).value = item
            #i = 0
            col = 2
            #while i < size:
            for sub in b[item]:
                sheet.cell(row=ro, column=col).value = sub
                #i+=1
                col+=1
            ro+=1
            
        sheet.cell(row=2, column=6).value = counter
    else:
        r = 2
        for item in b:
            sheet.cell(row=r, column=3).value = item
            r +=1

    
    wb.save("hospital.xlsx")    
        
        
        
        
        
        
        
        
        
        